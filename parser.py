"""Parse the two Excel budget files into structured JSON categories."""

from __future__ import annotations

import os
from dataclasses import dataclass, asdict
from typing import Sequence

import openpyxl

DATA_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "data")

LATESTMKT_FILE = None  # resolved at runtime
OAK_FILE = None


def _find_files() -> tuple[str | None, str | None]:
    """Find the two xlsx files in data/."""
    latestmkt = None
    oak = None
    for f in os.listdir(DATA_DIR):
        if not f.endswith(".xlsx"):
            continue
        lower = f.lower()
        if "obsah" in lower or "komunikace" in lower:
            oak = os.path.join(DATA_DIR, f)
        elif "mrktg" in lower or "rozpočet" in lower or "rozpocet" in lower:
            latestmkt = os.path.join(DATA_DIR, f)
    return latestmkt, oak


@dataclass(frozen=True)
class SourceItem:
    """A single line item parsed from an Excel source."""
    name: str
    amount: float
    category: str
    source: str
    description: str = ""
    owner: str = ""


def _cell_str(val: object) -> str:
    if val is None:
        return ""
    return str(val).strip()


def _cell_float(val: object) -> float:
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val)
    text = str(val).replace(",", "").replace(" ", "").strip()
    try:
        return float(text)
    except ValueError:
        return 0.0


def _is_header_or_total(name: str) -> bool:
    if not name:
        return True
    # Pure numbers are subtotals, not real items
    stripped = name.replace(" ", "").replace(",", "").replace(".", "")
    if stripped.isdigit():
        return True
    upper = name.upper()
    skip = [
        "CELKEM", "SOUČET", "NÁKLADY CELKEM", "FACELIFTY RESTAURACÍ",
        "INOVACE", "RETAIL", "UM -", "NOVÉ ZNAČKY", "FACELIFTY ZNAČKY",
        "REVIZE STRATEGIÍ", "BRANDOVÝ MARKETING", "SÓLOKAPR",
        "WEBY DEVELOPMENT", "WEBY SPRÁVA", "ROZVOJOVÝ MARKETING LOKÁL",
        "ROZVOJOVÝ MARKETING ARETAIL", "AMBIENTE FRANCHISE",
        "NÁKLAD", "PROFILY", "PERSONÁLNÍ", "VEDENÍ", "KVARTÁLNÍ",
        "LICENCE", "PODPORA ORGANICKÝCH", "SPRÁVA", "PRAVIDELNÁ",
        "ANALYTIKA", "PŘECHOD", "ROZVOJ ŠABLON", "INTERAKTIVNÍ",
        "KREATIVNÍ DODÁVKY",
    ]
    return any(kw in upper for kw in skip) or name.startswith("=")


# --- Main budget (latestmkt) parsers ---

def _parse_projekty(ws) -> list[SourceItem]:
    """Parse PROJEKTY RM with section-based categories.

    Section headers map to categories:
    - Rows before 'ROZVOJOVÝ MARKETING LOKÁL' -> RM Ambiente
    - 'ROZVOJOVÝ MARKETING LOKÁL' section -> RM Lokály
    - 'ROZVOJOVÝ MARKETING ARETAIL' section -> RM Aretail
    """
    SECTION_MARKERS = {
        "ROZVOJOVÝ MARKETING LOKÁL": "RM Lokály",
        "ROZVOJOVÝ MARKETING ARETAIL": "RM Aretail",
        "CELKEM ROZVOJOVÝ - AMBIENTE + LOKÁL": None,  # stop marker (full total row)
        "POZNÁMKY": None,  # stop marker
    }
    current_category = "RM Ambiente"
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 1:
            continue
        col_a = _cell_str(row[0])
        col_b = _cell_str(row[1]) if len(row) > 1 else ""
        # Check for section markers in col A
        if col_a:
            upper = col_a.upper()
            for marker, cat in SECTION_MARKERS.items():
                if marker in upper:
                    if cat is None:
                        current_category = ""
                    else:
                        current_category = cat
                    break
        if not current_category:
            continue
        if len(row) < 3:
            continue
        amount = _cell_float(row[2])
        if amount <= 0:
            continue
        # Item name: prefer col A, fall back to col B (sub-items)
        name = col_a if col_a else col_b
        if _is_header_or_total(name) or not name:
            continue
        owner = col_b if col_a else ""  # col B is owner when col A has name
        desc = _cell_str(row[3]) if len(row) > 3 else ""
        items.append(SourceItem(
            name=name, amount=amount, category=current_category,
            source="budget", description=desc, owner=owner,
        ))
    return items


def _parse_brand(ws) -> list[SourceItem]:
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 3:
            continue
        name = _cell_str(row[0])
        amount = _cell_float(row[2])
        if _is_header_or_total(name) or amount <= 0:
            continue
        owner = _cell_str(row[1]) if len(row) > 1 else ""
        desc = _cell_str(row[3]) if len(row) > 3 else ""
        items.append(SourceItem(
            name=name, amount=amount, category="Brand",
            source="budget", description=desc, owner=owner,
        ))
    return items


def _parse_weby(ws) -> list[SourceItem]:
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 3:
            continue
        name = _cell_str(row[0])
        amount = _cell_float(row[2])
        if _is_header_or_total(name) or amount <= 0:
            continue
        owner = _cell_str(row[1]) if len(row) > 1 else ""
        desc = _cell_str(row[3]) if len(row) > 3 else ""
        items.append(SourceItem(
            name=name, amount=amount, category="Weby",
            source="budget", description=desc, owner=owner,
        ))
    return items


def _parse_vp(ws) -> list[SourceItem]:
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 14:
            continue
        name = _cell_str(row[1])
        if _is_header_or_total(name) or not name:
            continue
        marker = _cell_str(row[0])
        if marker.lower() == "x":
            continue
        annual = sum(_cell_float(row[i]) for i in range(2, 14))
        if annual <= 0:
            continue
        items.append(SourceItem(
            name=name, amount=annual, category="Věrnostní program",
            source="budget",
        ))
    return items


def _parse_ambi(ws) -> list[SourceItem]:
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 8:
            continue
        name = _cell_str(row[0])
        if not name or _is_header_or_total(name):
            continue
        amount = _cell_float(row[7])
        if amount <= 0:
            amount = _cell_float(row[6])
        if amount <= 0:
            continue
        desc = _cell_str(row[1])
        items.append(SourceItem(
            name=name, amount=amount, category="Provozní (AMBI)",
            source="budget", description=desc,
        ))
    return items


def _parse_af(ws) -> list[SourceItem]:
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 3:
            continue
        name = _cell_str(row[0])
        amount = _cell_float(row[2])
        if _is_header_or_total(name) or amount <= 0:
            continue
        owner = _cell_str(row[1]) if len(row) > 1 else ""
        desc = _cell_str(row[3]) if len(row) > 3 else ""
        items.append(SourceItem(
            name=name, amount=amount, category="AF",
            source="budget", description=desc, owner=owner,
        ))
    return items


# --- OaK parsers ---

def _parse_oak_detail(ws, sheet_name: str) -> list[SourceItem]:
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 2:
            continue
        name = _cell_str(row[0])
        if not name or _is_header_or_total(name):
            continue
        annual = 0.0
        if len(row) > 13:
            annual = _cell_float(row[13])
        if annual <= 0:
            annual = sum(_cell_float(row[i]) for i in range(1, min(13, len(row))))
        if annual <= 0:
            continue
        items.append(SourceItem(
            name=name, amount=annual, category=sheet_name,
            source="oak",
        ))
    return items


def _parse_oak_lide(ws) -> list[SourceItem]:
    # Subtotal/section rows to skip — these are area summaries, not people
    LIDE_SKIP = {
        "PR A OBSAH", "MAILING", "SOCIAL", "ACCOUNT",
        "VEDENI A ADMIN", "VEDENÍ A ADMIN",
    }
    items = []
    for row in ws.iter_rows(min_row=5, max_row=ws.max_row, values_only=True):
        if len(row) < 2:
            continue
        name = _cell_str(row[0])
        if not name or _is_header_or_total(name):
            continue
        upper = name.upper()
        if any(kw in upper for kw in ["NÁKLADY", "SPOLEČNÍ", "NÁBOR"]):
            continue
        if upper in LIDE_SKIP:
            continue
        annual = 0.0
        if len(row) > 13:
            annual = _cell_float(row[13])
        if annual <= 0:
            annual = sum(_cell_float(row[i]) for i in range(1, min(13, len(row))))
        if annual <= 0:
            continue
        area = _cell_str(row[14]) if len(row) > 14 else ""
        items.append(SourceItem(
            name=name, amount=annual, category="OaK Lidé",
            source="oak", description=f"oblast: {area}" if area else "",
        ))
    return items


def _parse_oak_aktivity(ws) -> list[SourceItem]:
    items = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if len(row) < 2:
            continue
        name = _cell_str(row[0])
        amount = _cell_float(row[1])
        if not name or amount <= 0 or _is_header_or_total(name):
            continue
        items.append(SourceItem(
            name=name, amount=amount, category="OaK Aktivity",
            source="oak",
        ))
    return items


# --- Main parse functions ---

def parse_latestmkt(path: str) -> list[SourceItem]:
    wb = openpyxl.load_workbook(path, data_only=True)
    items: list[SourceItem] = []
    parsers = {
        "PROJEKTY RM": _parse_projekty,
        "BRANDOVÝ MRKTG.": _parse_brand,
        "WEBY": _parse_weby,
        "VP": _parse_vp,
        "Co zůstalo v AMBI": _parse_ambi,
        "AF": _parse_af,
    }
    for sheet_name, fn in parsers.items():
        if sheet_name in wb.sheetnames:
            items.extend(fn(wb[sheet_name]))
    return items


def parse_oak(path: str) -> tuple[list[SourceItem], float]:
    """Parse OaK file using Celkem sheet as source of truth.

    Returns (items, official_total).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    items: list[SourceItem] = []
    official_total = 0.0

    # Parse Celkem sheet — rows 3-19 have the real structure
    # Row 3: grand total, rows 4+ are categories with sub-rows
    if "Celkem" not in wb.sheetnames:
        return items, 0.0

    ws = wb["Celkem"]
    # Row 3 = grand total
    for row in ws.iter_rows(min_row=3, max_row=3, values_only=True):
        official_total = _cell_float(row[1])
        break

    # Rows 4-19: main categories (bold) and sub-rows (Externí náklady, Lidé)
    # Main categories: PR a obsah, Sociální sítě, Email marketing,
    #   Account mngmnt, Vedení a administrativa dodávky, Rezerva
    MAIN_CATEGORIES = {
        "PR a obsah", "Sociální sítě", "Email marketing",
        "Account mngmnt", "Vedení a administrativa dodávky", "Rezerva",
    }
    SUB_ROWS = {"Externí náklady", "Lidé"}

    current_cat = ""
    current_subs: list[str] = []
    for row in ws.iter_rows(min_row=4, max_row=19, values_only=True):
        name = _cell_str(row[0])
        val = _cell_float(row[1])
        if not name:
            continue
        if name in MAIN_CATEGORIES:
            # Emit previous category if any
            if current_cat:
                pass  # already emitted
            current_cat = name
            current_subs = []
            items.append(SourceItem(
                name=name, amount=val, category="OaK",
                source="oak",
            ))
        elif name in SUB_ROWS and current_cat and val > 0:
            current_subs.append(f"{name}: {val:,.0f}")

    # Also include aktivity k jednání
    if "aktivity k jednání" in wb.sheetnames:
        items.extend(_parse_oak_aktivity(wb["aktivity k jednání"]))

    return items, official_total


def parse_all() -> dict:
    """Parse both files, return structured data grouped by category."""
    latestmkt_path, oak_path = _find_files()

    budget_items: list[SourceItem] = []
    oak_items: list[SourceItem] = []
    oak_official_total = 0.0

    if latestmkt_path:
        budget_items = parse_latestmkt(latestmkt_path)
    if oak_path:
        oak_items, oak_official_total = parse_oak(oak_path)

    # Group by category
    categories: dict[str, list[dict]] = {}
    for item in budget_items:
        cat = item.category
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(asdict(item))

    oak_categories: dict[str, list[dict]] = {}
    for item in oak_items:
        cat = item.category
        if cat not in oak_categories:
            oak_categories[cat] = []
        oak_categories[cat].append(asdict(item))

    budget_total = sum(i.amount for i in budget_items)

    # Get overall total from CELKEM MRKTG sheet (row 4, col B = NÁKLADY CELKEM)
    overall_total = budget_total + oak_official_total  # fallback
    if latestmkt_path:
        wb = openpyxl.load_workbook(latestmkt_path, data_only=True)
        if "CELKEM MRKTG" in wb.sheetnames:
            ws = wb["CELKEM MRKTG"]
            for row in ws.iter_rows(min_row=4, max_row=4, values_only=True):
                val = _cell_float(row[1])
                if val > 0:
                    overall_total = val
                break

    return {
        "budget_categories": categories,
        "oak_categories": oak_categories,
        "budget_total": budget_total,
        "oak_total": oak_official_total,
        "overall_total": overall_total,
    }
